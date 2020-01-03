import os, openpyxl
import pandas as pd
import re
from openpyxl.utils.dataframe import dataframe_to_rows

#описываем ф-цию формирования словарей опеторов и товаров из соответствующих файлов
def to_dict(file, out):
    with open(file, encoding='utf-8') as f:
        i = 1
        for line in f:
            line = line.rstrip('\n')
            out[line.split(':')[0]]=line.split(':')[1]
            i += 1

      
dict_offers={}
dict_operators={}

#формируем словари с "правильными" именами товаров и операторов для корректного поиска по файлу-отчету
to_dict('data\\offers.txt', dict_offers)
to_dict('data\\operators.txt', dict_operators)

cwd = os.getcwd()
os.chdir('income\\')
listDir = os.listdir()


data_new = []

##перебираем поочереди файлы в директории "входа" и считываем строки с них
for files in listDir:
    wb = openpyxl.load_workbook(files)
    sheet = wb.active
    df = pd.DataFrame(sheet.values)
    df_len = len(df)
    for num in range(0, df_len):
        l = []
        #делаем подмену имен операторов и товаров из соответствующих словарей
        for row in df.loc[num]:
            for operator in dict_operators:
                if row == operator:
                    row = dict_operators[operator]
            for offer in dict_offers:
                if row == offer:
                    row = dict_offers[offer]
            l.append(row)
        if l[0] != 'offer_name':
            l = tuple(l)
            data_new.append(l)
        
os.chdir(cwd)
os.chdir('out\\')

#открываем файл для ввода данных из выгрузки
wb2 = openpyxl.load_workbook('out.xlsx')
ws2 = wb2.active
cout_row = len(data_new)

data_added = [] #создаем пустой список для успешно добавленых строк
count_added = 0
for line in range (1, cout_row+1):
    count_line = 0
    for row in ws2.rows:
        count_line += 1
        if (str(row[0].value) == data_new[line-1][1]) and (str(row[1].value) == data_new[line-1][0]):
            ws2.cell(row=count_line, column=6, value = data_new[line-1][3])
            ws2.cell(row=count_line, column=9, value = data_new[line-1][2])
            data_added.append(data_new[line-1])
            count_added += 1
            
wb2.save('out_new.xlsx')

#проверяем все ли строки добавились в файл-отчет
result = list(set(data_new) ^ set(data_added))

#выводим те строки, которые не добавились по каким-либо причинам
for no_added in result:
    print(no_added)
print('==================================')
print('Работа скрипта завершена!')
